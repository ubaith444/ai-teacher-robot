"""
app/utils/helpers.py
────────────────────
General-purpose utility functions for the Zoro Robot Attendance System.
No external I/O — pure Python helpers only.

Contents
--------
• Session ID generation + parsing
• Time / timezone helpers (defaults to IST for Indian school deployments)
• CSV and Excel export builders
• JSON-safe helpers
• Pagination metadata
"""

from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime, time, timedelta
from typing import Any, Dict, Iterable, List, Optional
from zoneinfo import ZoneInfo

import structlog

log = structlog.get_logger(__name__)


# ── Session ID generation ──────────────────────────────────────────────────────


def build_session_id(target_date: date, class_section: str, period_number: int) -> str:
    """
    Build a deterministic, human-readable session identifier.

    Format: '<YYYY-MM-DD>_<class_section>_P<period_number>'
    Example: '2024-06-01_10-A_P3'
    """
    return f"{target_date.isoformat()}_{class_section}_P{period_number}"


def parse_session_id(session_id: str) -> Optional[Dict[str, Any]]:
    """
    Parse a session_id produced by build_session_id().

    Returns a dict: {date, class_section, period_number} or None on failure.
    Handles class_sections that contain underscores (e.g. '10_A').
    """
    try:
        parts = session_id.split("_")
        date_str = parts[0]
        period_str = parts[-1]  # 'P3'
        class_section = "_".join(parts[1:-1])  # everything between date and period
        return {
            "date": date.fromisoformat(date_str),
            "class_section": class_section,
            "period_number": int(period_str.lstrip("P")),
        }
    except (IndexError, ValueError):
        return None


# ── Time / timezone helpers ────────────────────────────────────────────────────


def _tz(tz_name: str) -> ZoneInfo:
    """Return ZoneInfo for the given IANA timezone name."""
    return ZoneInfo(tz_name)


def now_local(tz_name: str = "Asia/Kolkata") -> datetime:
    """Return the current datetime in the specified local timezone (default IST)."""
    return datetime.now(tz=_tz(tz_name))


def current_time_local(tz_name: str = "Asia/Kolkata") -> time:
    """Return the current wall-clock time (without timezone) in local tz."""
    return now_local(tz_name).replace(tzinfo=None).time()


def current_date_local(tz_name: str = "Asia/Kolkata") -> date:
    """Return today's date in local timezone."""
    return now_local(tz_name).date()


def day_of_week_str(
    target_date: Optional[date] = None, tz_name: str = "Asia/Kolkata"
) -> str:
    """
    Return the lowercase English day-of-week string.
    Defaults to today in local timezone.
    """
    d = target_date or current_date_local(tz_name)
    return d.strftime("%A").lower()


def time_is_within(t: time, start: time, end: time) -> bool:
    """Return True if t is in [start, end)."""
    return start <= t < end


def compute_attendance_status(session_start: Optional[time], grace_minutes: int) -> str:
    """
    Return 'present' or 'late' based on the current local time, the session
    start time, and the configured grace period.

    If session_start is None → default to 'present'.
    """
    if session_start is None:
        return "present"
    dummy = datetime(2000, 1, 1)
    start_dt = datetime.combine(dummy, session_start)
    grace_end = start_dt + timedelta(minutes=grace_minutes)
    now_dt = datetime.combine(dummy, current_time_local())
    return "present" if now_dt <= grace_end else "late"


# ── CSV export ─────────────────────────────────────────────────────────────────


def build_csv_bytes(rows: Iterable[Dict[str, Any]], fieldnames: List[str]) -> bytes:
    """
    Serialise rows to a UTF-8 BOM CSV byte string (Excel-friendly encoding).

    Args:
        rows:       Iterable of dicts; missing keys → empty string.
        fieldnames: Column order (also the header row).

    Returns:
        bytes of the complete CSV file with UTF-8 BOM.
    """
    buf = io.StringIO()
    writer = csv.DictWriter(
        buf,
        fieldnames=fieldnames,
        extrasaction="ignore",
        lineterminator="\r\n",
    )
    writer.writeheader()
    for row in rows:
        writer.writerow({k: row.get(k, "") for k in fieldnames})
    # Prepend UTF-8 BOM so Excel auto-detects encoding
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


# ── Excel export ───────────────────────────────────────────────────────────────


def build_excel_bytes(rows: Iterable[Dict[str, Any]], fieldnames: List[str]) -> bytes:
    """
    Serialise rows to an in-memory .xlsx file using openpyxl.

    Applies bold formatting to the header row and auto-fits column widths
    for readability.

    Raises:
        RuntimeError: if openpyxl is not installed.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header with bold font
    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Data rows
    row_list = list(rows)
    for row in row_list:
        ws.append(
            [str(row.get(f, "")) if row.get(f) is not None else "" for f in fieldnames]
        )

    # Auto-fit column widths (approximate)
    for col_idx, field in enumerate(fieldnames, start=1):
        max_len = max(
            len(str(field)),
            *(len(str(r.get(field, ""))) for r in row_list),
            default=len(field),
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_len + 2, 50
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── JSON helpers ───────────────────────────────────────────────────────────────


def safe_json_loads(raw: Any, fallback: Any = None) -> Any:
    """
    Try to JSON-decode *raw*; return *fallback* on any failure.
    Already-decoded dicts/lists are returned as-is.
    """
    if raw is None:
        return fallback
    if isinstance(raw, (dict, list)):
        return raw  # already decoded by SQLAlchemy JSONB
    try:
        return json.loads(raw)
    except (TypeError, json.JSONDecodeError):
        return fallback


# ── Pagination ─────────────────────────────────────────────────────────────────


def paginate_meta(total: int, page: int, page_size: int) -> Dict[str, int]:
    """Return standard pagination metadata dict."""
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": max(1, (total + page_size - 1) // page_size),
    }


"""
Zoro AI Robot - General utility helpers.
"""

import logging
import time
from functools import wraps
from typing import Callable

logger = logging.getLogger("zoro.utils")


def clamp(value: int, min_val: int, max_val: int) -> int:
    """Clamp an integer between min and max inclusive."""
    return max(min_val, min(max_val, value))


def rate_limit_check(last_call: float, min_interval: float) -> bool:
    """Return True if enough time has passed since last_call."""
    return (time.monotonic() - last_call) >= min_interval


def async_retry(retries: int = 2, delay: float = 0.5):
    """
    Decorator: retry an async function up to `retries` times on exception.
    Useful for flaky service calls on the Pi's network.
    """

    def decorator(func: Callable):
        @wraps(func)
        async def wrapper(*args, **kwargs):
            import asyncio

            last_exc = None
            for attempt in range(retries + 1):
                try:
                    return await func(*args, **kwargs)
                except Exception as exc:
                    last_exc = exc
                    if attempt < retries:
                        logger.warning(
                            f"Retry {attempt + 1}/{retries} for {func.__name__}: {exc}"
                        )
                        await asyncio.sleep(delay)
            raise last_exc

        return wrapper

    return decorator


def truncate(text: str, max_chars: int = 200) -> str:
    """Truncate a string for safe logging."""
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "…"

